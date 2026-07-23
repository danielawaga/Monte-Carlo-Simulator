"""Generate a fictitious correlated Excel risk register example."""

from pathlib import Path

from openpyxl import Workbook

from monte_carlo_simulator.io.schema import METADATA_KEYS, RISK_REGISTER_COLUMNS

ROWS = [
    {
        "name": "Design",
        "distribution": "triangular",
        "minimum": 80,
        "most_likely": 100,
        "maximum": 140,
        "category": "engineering",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Prototype",
        "distribution": "pert",
        "minimum": 120,
        "most_likely": 160,
        "maximum": 240,
        "lambda_shape": 4,
        "category": "engineering",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Testing",
        "distribution": "uniform",
        "minimum": 50,
        "maximum": 90,
        "category": "quality",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Permits",
        "distribution": "normal",
        "mean": 70,
        "standard_deviation": 10,
        "category": "admin",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Supplier",
        "distribution": "lognormal",
        "mean": 200,
        "standard_deviation": 45,
        "category": "procurement",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Bonus",
        "distribution": "event",
        "probability": 0.25,
        "impact": -30,
        "category": "commercial",
        "unit": "kEUR",
        "enabled": True,
    },
    {
        "name": "Penalty",
        "distribution": "event",
        "probability": 0.15,
        "impact": 60,
        "category": "commercial",
        "unit": "kEUR",
        "enabled": True,
    },
]
CORR = [
    [1, 0.35, 0.20, 0.10, 0.25, -0.20, 0.15],
    [0.35, 1, 0.30, 0.15, 0.20, -0.10, 0.20],
    [0.20, 0.30, 1, 0.25, 0.10, -0.15, 0.30],
    [0.10, 0.15, 0.25, 1, 0.05, 0, 0.10],
    [0.25, 0.20, 0.10, 0.05, 1, -0.10, 0.25],
    [-0.20, -0.10, -0.15, 0, -0.10, 1, -0.25],
    [0.15, 0.20, 0.30, 0.10, 0.25, -0.25, 1],
]
META = {
    "schema_version": "1.0",
    "project_name": "Fictitious correlated cost register",
    "analysis_type": "cost",
    "default_unit": "kEUR",
    "baseline_estimate": 900,
    "description": "Synthetic public example; no client or real project data.",
}


def generate(path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    metadata_sheet = workbook.active
    metadata_sheet.title = "metadata"
    metadata_sheet.append(["key", "value"])
    for key in METADATA_KEYS:
        metadata_sheet.append([key, META.get(key)])

    register_sheet = workbook.create_sheet("risk_register")
    register_sheet.append(list(RISK_REGISTER_COLUMNS))
    for row in ROWS:
        register_sheet.append([row.get(column) for column in RISK_REGISTER_COLUMNS])

    correlation_sheet = workbook.create_sheet("correlations")
    names = [row["name"] for row in ROWS]
    correlation_sheet.append([None, *names])
    for name, coefficients in zip(names, CORR, strict=True):
        correlation_sheet.append([name, *coefficients])

    workbook.save(path)
    workbook.close()
    return path


if __name__ == "__main__":
    generate(Path("data/examples/correlated_cost_register.xlsx"))
