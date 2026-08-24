# ruff: noqa: E501
"""Build the synthetic building risk register from the public schema-v1 template."""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
# Sorties de travail de l'étude de cas : générées, jamais versionnées.
WORK = ROOT / "data" / "output" / "case_study"
TEMPLATE = ROOT / "data" / "templates" / "risk_register_template.xlsx"
OUTPUT = WORK / "registre_risques_batiment_synthetique.xlsx"


ROWS = [
    [
        "Études et contrôles",
        "normal",
        None,
        None,
        None,
        250_000,
        20_000,
        None,
        None,
        None,
        "Études",
        None,
        True,
        "Hypothèse synthétique pédagogique ; prestation relativement prévisible.",
    ],
    [
        "Installation du chantier",
        "pert",
        220_000,
        280_000,
        380_000,
        None,
        None,
        None,
        None,
        None,
        "Préparation",
        None,
        True,
        "Hypothèse synthétique pédagogique ; dépend de la durée et des moyens temporaires.",
    ],
    [
        "Terrassement",
        "pert",
        300_000,
        400_000,
        650_000,
        None,
        None,
        None,
        None,
        None,
        "Gros œuvre",
        None,
        True,
        "Hypothèse synthétique pédagogique ; incertitude sur le sol, les volumes et les évacuations.",
    ],
    [
        "Fondations et gros œuvre",
        "pert",
        1_600_000,
        1_900_000,
        2_500_000,
        None,
        None,
        None,
        None,
        None,
        "Gros œuvre",
        None,
        True,
        "Hypothèse synthétique pédagogique ; exposition aux quantités de béton et d’acier.",
    ],
    [
        "Étanchéité et toiture",
        "triangular",
        400_000,
        500_000,
        700_000,
        None,
        None,
        None,
        None,
        None,
        "Enveloppe",
        None,
        True,
        "Hypothèse synthétique pédagogique ; détails techniques susceptibles d’évoluer.",
    ],
    [
        "Façades et menuiseries",
        "pert",
        650_000,
        750_000,
        1_000_000,
        None,
        None,
        None,
        None,
        None,
        "Enveloppe",
        None,
        True,
        "Hypothèse synthétique pédagogique ; sensible aux matériaux, vitrages et fournisseurs.",
    ],
    [
        "Électricité",
        "pert",
        550_000,
        650_000,
        850_000,
        None,
        None,
        None,
        None,
        None,
        "Lots techniques",
        None,
        True,
        "Hypothèse synthétique pédagogique ; dépend du niveau d’équipement et des réseaux.",
    ],
    [
        "Plomberie et sanitaires",
        "pert",
        400_000,
        480_000,
        650_000,
        None,
        None,
        None,
        None,
        None,
        "Lots techniques",
        None,
        True,
        "Hypothèse synthétique pédagogique ; incertitude sur les quantités et les gammes.",
    ],
    [
        "Climatisation et ventilation",
        "pert",
        450_000,
        550_000,
        800_000,
        None,
        None,
        None,
        None,
        None,
        "Lots techniques",
        None,
        True,
        "Hypothèse synthétique pédagogique ; exposition au dimensionnement et aux prix.",
    ],
    [
        "Revêtements et finitions",
        "pert",
        800_000,
        950_000,
        1_300_000,
        None,
        None,
        None,
        None,
        None,
        "Finitions",
        None,
        True,
        "Hypothèse synthétique pédagogique ; choix tardifs, reprises et exigences esthétiques.",
    ],
    [
        "Aménagements extérieurs",
        "pert",
        350_000,
        450_000,
        700_000,
        None,
        None,
        None,
        None,
        None,
        "Extérieurs",
        None,
        True,
        "Hypothèse synthétique pédagogique ; périmètre moins défini et dépendant du site.",
    ],
    [
        "Supervision du chantier",
        "normal",
        None,
        None,
        None,
        300_000,
        45_000,
        None,
        None,
        None,
        "Pilotage",
        None,
        True,
        "Hypothèse synthétique pédagogique ; coût régulier mais sensible aux prolongations.",
    ],
    [
        "Sol plus difficile que prévu",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.12,
        350_000,
        None,
        "Risque géotechnique",
        None,
        True,
        "Hypothèse synthétique pédagogique ; menace liée à des reconnaissances incomplètes.",
    ],
    [
        "Hausse exceptionnelle des matériaux",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.20,
        450_000,
        None,
        "Risque marché",
        None,
        True,
        "Hypothèse synthétique pédagogique ; choc possible sur acier, ciment et équipements.",
    ],
    [
        "Retard administratif ou de raccordement",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.18,
        250_000,
        None,
        "Risque planning",
        None,
        True,
        "Hypothèse synthétique pédagogique ; peut prolonger la mobilisation du chantier.",
    ],
    [
        "Modification tardive du programme",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.15,
        300_000,
        None,
        "Risque programme",
        None,
        True,
        "Hypothèse synthétique pédagogique ; études, reprises ou commandes supplémentaires.",
    ],
    [
        "Défaillance ou remplacement d’un fournisseur",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.10,
        220_000,
        None,
        "Risque fournisseur",
        None,
        True,
        "Hypothèse synthétique pédagogique ; nouvel achat, écart de prix ou retard.",
    ],
    [
        "Négociation commerciale favorable",
        "event",
        None,
        None,
        None,
        None,
        None,
        0.25,
        -120_000,
        None,
        "Opportunité",
        None,
        True,
        "Hypothèse synthétique pédagogique ; remise ou regroupement de commandes.",
    ],
]


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, 15):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(TEMPLATE)

    metadata = workbook["metadata"]
    metadata_values = {
        "schema_version": "1.0",
        "project_name": "Bâtiment administratif R+1 — cas synthétique",
        "analysis_type": "cost",
        "default_unit": "MAD",
        "baseline_estimate": 7_750_000,
        "description": "Cas d’étude synthétique réaliste et pédagogique ; montants HT, hors terrain, financement et fiscalité.",
    }
    for row in range(2, metadata.max_row + 1):
        key = metadata.cell(row, 1).value
        if key in metadata_values:
            metadata.cell(row, 2).value = metadata_values[key]
    metadata["B6"].number_format = '#,##0 "MAD"'

    register = workbook["risk_register"]
    source_style_row = 2
    for row in range(2, max(register.max_row, len(ROWS) + 1) + 1):
        if row > 2:
            copy_row_style(register, source_style_row, row)
        for column in range(1, 15):
            register.cell(row, column).value = None

    for row_index, values in enumerate(ROWS, start=2):
        for column_index, value in enumerate(values, start=1):
            register.cell(row_index, column_index).value = value
        for column in range(3, 11):
            register.cell(row_index, column).number_format = "#,##0.00"
        register.cell(row_index, 8).number_format = "0.00%"

    register.auto_filter.ref = f"A1:N{len(ROWS) + 1}"
    register.sheet_view.showGridLines = False
    metadata.sheet_view.showGridLines = False
    workbook["instructions"].sheet_view.showGridLines = False

    workbook.properties.title = "Registre de risques — bâtiment administratif synthétique"
    workbook.properties.subject = "Cas d’étude de coût, schéma Excel 1.0"
    workbook.properties.description = "Données entièrement synthétiques et pédagogiques."
    workbook.save(OUTPUT)
    workbook.close()
    print(OUTPUT)


if __name__ == "__main__":
    main()
