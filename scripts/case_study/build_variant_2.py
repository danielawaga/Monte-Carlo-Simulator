# ruff: noqa: E501
"""Create variant 2 of the synthetic building register with moderate correlations."""

from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
# Sorties de travail de l'étude de cas : générées, jamais versionnées.
WORK = ROOT / "data" / "output" / "case_study"
SOURCE = WORK / "registre_risques_batiment_synthetique.xlsx"
OUTPUT = WORK / "registre_risques_batiment_correlations.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DIAGONAL_FILL = PatternFill("solid", fgColor="D9EAF7")


PAIRS = {
    frozenset(("Terrassement", "Fondations et gros œuvre")): (
        0.30,
        "Conditions de sol défavorables susceptibles d’affecter les terrassements et les fondations.",
    ),
    frozenset(("Fondations et gros œuvre", "Hausse exceptionnelle des matériaux")): (
        0.50,
        "Le gros œuvre consomme beaucoup de béton et d’acier et est exposé au choc de prix.",
    ),
    frozenset(("Électricité", "Plomberie et sanitaires")): (
        0.25,
        "Les modifications d’aménagement peuvent affecter simultanément plusieurs réseaux.",
    ),
    frozenset(("Supervision du chantier", "Retard administratif ou de raccordement")): (
        0.40,
        "Un retard peut prolonger la durée de mobilisation de la supervision.",
    ),
}


def main() -> None:
    workbook = load_workbook(SOURCE)
    if "correlations" in workbook.sheetnames:
        del workbook["correlations"]

    metadata = workbook["metadata"]
    metadata["B3"] = "Bâtiment administratif R+1 — variante 2 corrélée"
    metadata["B7"] = (
        "Cas synthétique pédagogique avec quatre corrélations modérées et justifiées ; "
        "mêmes coûts marginaux et même baseline que la variante 1."
    )

    register = workbook["risk_register"]
    names = [
        register.cell(row, 1).value
        for row in range(2, register.max_row + 1)
        if register.cell(row, 13).value is not False
    ]
    if not names or any(not isinstance(name, str) for name in names):
        raise ValueError("Active risk item names could not be read from the source workbook.")

    matrix = np.eye(len(names), dtype=float)
    index = {name: position for position, name in enumerate(names)}
    for pair, (coefficient, _reason) in PAIRS.items():
        first, second = tuple(pair)
        i, j = index[first], index[second]
        matrix[i, j] = coefficient
        matrix[j, i] = coefficient

    eigenvalues = np.linalg.eigvalsh(matrix)
    if eigenvalues.min() <= 0:
        raise ValueError("The proposed correlation matrix is not strictly positive definite.")

    sheet = workbook.create_sheet("correlations")
    sheet.append([None, *names])
    for row_name, values in zip(names, matrix, strict=True):
        sheet.append([row_name, *values.tolist()])

    sheet.freeze_panes = "B2"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 46
    sheet.row_dimensions[1].height = 150

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            text_rotation=90, horizontal="center", vertical="bottom", wrap_text=True
        )
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, len(names) + 2):
        name_cell = sheet.cell(row, 1)
        name_cell.fill = HEADER_FILL
        name_cell.font = HEADER_FONT
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 31
        for column in range(2, len(names) + 2):
            cell = sheet.cell(row, column)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 5.5
            if row - 1 == column - 1:
                cell.fill = DIAGONAL_FILL

    last_cell = sheet.cell(len(names) + 1, len(names) + 1).coordinate
    sheet.conditional_formatting.add(
        f"B2:{last_cell}",
        ColorScaleRule(
            start_type="num",
            start_value=-1,
            start_color="F8696B",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )

    for pair, (coefficient, reason) in PAIRS.items():
        first, second = tuple(pair)
        for row_name, column_name in ((first, second), (second, first)):
            row = index[row_name] + 2
            column = index[column_name] + 2
            sheet.cell(row, column).comment = Comment(
                f"Coefficient proposé : {coefficient:.2f}. {reason} Valeur à valider avec un expert métier.",
                "User",
            )

    sheet.auto_filter.ref = f"A1:{last_cell}"
    workbook.properties.title = "Registre de risques — bâtiment administratif corrélé"
    workbook.properties.description = (
        "Variante 2 du cas synthétique avec quatre corrélations modérées."
    )
    workbook.save(OUTPUT)
    workbook.close()
    print(OUTPUT)
    print(f"minimum_eigenvalue={eigenvalues.min():.6f}")


if __name__ == "__main__":
    main()
