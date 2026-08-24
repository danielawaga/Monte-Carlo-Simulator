# ruff: noqa: E501
"""Create variant 3 of the synthetic building register with targeted stress."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
# Sorties de travail de l'étude de cas : générées, jamais versionnées.
WORK = ROOT / "data" / "output" / "case_study"
SOURCE = WORK / "registre_risques_batiment_correlations.xlsx"
OUTPUT = WORK / "registre_risques_batiment_stress.xlsx"

CONTINUOUS_STRESS = {
    "Terrassement": {"most_likely": 430_000, "maximum": 750_000},
    "Fondations et gros œuvre": {"most_likely": 2_000_000, "maximum": 2_800_000},
    "Façades et menuiseries": {"maximum": 1_100_000},
    "Climatisation et ventilation": {"most_likely": 580_000, "maximum": 900_000},
    "Revêtements et finitions": {"most_likely": 1_000_000, "maximum": 1_500_000},
    "Aménagements extérieurs": {"maximum": 800_000},
    "Supervision du chantier": {"mean": 320_000, "standard_deviation": 60_000},
}

EVENT_STRESS = {
    "Sol plus difficile que prévu": {"probability": 0.18, "impact": 450_000},
    "Hausse exceptionnelle des matériaux": {"probability": 0.30, "impact": 600_000},
    "Retard administratif ou de raccordement": {"probability": 0.25, "impact": 350_000},
    "Modification tardive du programme": {"probability": 0.20, "impact": 400_000},
    "Défaillance ou remplacement d’un fournisseur": {"probability": 0.15, "impact": 300_000},
    "Négociation commerciale favorable": {"probability": 0.15, "impact": -80_000},
}


def main() -> None:
    workbook = load_workbook(SOURCE)
    metadata = workbook["metadata"]
    metadata["B3"] = "Bâtiment administratif R+1 — variante 3 stressée"
    metadata["B7"] = (
        "Cas synthétique pédagogique : corrélations de la variante 2 conservées et stress ciblé "
        "sur les lots exposés, les menaces et l’opportunité commerciale ; baseline inchangée."
    )
    metadata.column_dimensions["B"].width = 68
    metadata.column_dimensions["C"].width = 72
    metadata["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    metadata["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    metadata.row_dimensions[3].height = 30
    metadata.row_dimensions[7].height = 44

    register = workbook["risk_register"]
    headers = {cell.value: cell.column for cell in register[1]}
    changes = []
    for row in range(2, register.max_row + 1):
        name = register.cell(row, headers["name"]).value
        updates = CONTINUOUS_STRESS.get(name) or EVENT_STRESS.get(name)
        if not updates:
            continue
        descriptions = []
        for field, new_value in updates.items():
            cell = register.cell(row, headers[field])
            old_value = cell.value
            cell.value = new_value
            descriptions.append(f"{field}: {old_value} → {new_value}")
            changes.append((name, field, old_value, new_value))
        note_cell = register.cell(row, headers["notes"])
        prefix = f"{note_cell.value} " if note_cell.value else ""
        note_cell.value = prefix + "Variante 3 - stress ciblé : " + "; ".join(descriptions) + "."

    if "stress_changes" in workbook.sheetnames:
        del workbook["stress_changes"]
    sheet = workbook.create_sheet("stress_changes", 2)
    sheet.append(["Poste", "Paramètre", "Variante 2", "Variante 3", "Justification"])
    reasons = {
        "Terrassement": "Contexte géotechnique plus défavorable.",
        "Fondations et gros œuvre": "Quantités et prix des matériaux sous tension.",
        "Façades et menuiseries": "Risque fournisseur et spécifications renforcées.",
        "Climatisation et ventilation": "Équipements importés et redimensionnement possible.",
        "Revêtements et finitions": "Choix tardifs et reprises plus probables.",
        "Aménagements extérieurs": "Périmètre moins mature.",
        "Supervision du chantier": "Durée de mobilisation accrue et plus variable.",
        "Sol plus difficile que prévu": "Reconnaissances incomplètes en scénario défavorable.",
        "Hausse exceptionnelle des matériaux": "Choc de prix plus fréquent et plus sévère.",
        "Retard administratif ou de raccordement": "Délais externes renforcés.",
        "Modification tardive du programme": "Arbitrages tardifs plus probables.",
        "Défaillance ou remplacement d’un fournisseur": "Chaîne d’approvisionnement fragilisée.",
        "Négociation commerciale favorable": "Opportunité moins probable et moins avantageuse.",
    }
    for name, field, old, new in changes:
        sheet.append([name, field, old, new, reasons[name]])

    navy = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    widths = {"A": 46, "B": 24, "C": 18, "D": 18, "E": 58}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 32
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row, 3).number_format = (
            "0.00%" if sheet.cell(row, 2).value == "probability" else "#,##0"
        )
        sheet.cell(row, 4).number_format = (
            "0.00%" if sheet.cell(row, 2).value == "probability" else "#,##0"
        )

    workbook.properties.title = "Registre de risques — bâtiment administratif stressé"
    workbook.properties.description = (
        "Variante 3 avec corrélations conservées et stress ciblé documenté."
    )
    workbook.save(OUTPUT)
    workbook.close()
    print(OUTPUT)
    print(f"changes={len(changes)}")


if __name__ == "__main__":
    main()
