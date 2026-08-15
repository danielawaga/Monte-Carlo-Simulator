from pathlib import Path

from openpyxl import load_workbook

from monte_carlo_simulator.io import create_risk_register_template, create_risk_register_workbook


def test_template_contains_documentation_and_excel_validations(tmp_path: Path) -> None:
    path = create_risk_register_template(tmp_path / "template.xlsx")

    workbook = load_workbook(path)
    try:
        assert workbook.sheetnames == ["metadata", "risk_register", "instructions"]
        assert workbook["metadata"]["B2"].value == "1.0"
        assert workbook["risk_register"].freeze_panes == "A2"
        assert workbook["risk_register"].auto_filter.ref == "A1:N7"
        assert len(workbook["risk_register"].data_validations.dataValidation) == 2
        instructions = workbook["instructions"]
        assert any(
            "Never commit real client" in str(cell.value)
            for row in instructions.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()


def test_consultant_workbook_contains_entered_hypotheses(tmp_path: Path) -> None:
    path = create_risk_register_workbook(
        tmp_path / "consultant.xlsx",
        metadata_values={
            "project_name": "Projet consultant",
            "analysis_type": "cost",
            "default_unit": "MAD",
            "baseline_estimate": 125_000,
            "description": "Hypothèses préparées dans l'interface.",
        },
        risk_rows=[
            {
                "name": "Études",
                "distribution": "triangular",
                "minimum": 10_000,
                "most_likely": 15_000,
                "maximum": 25_000,
                "category": "Ingénierie",
                "enabled": True,
            }
        ],
    )

    workbook = load_workbook(path)
    try:
        assert workbook["metadata"]["B3"].value == "Projet consultant"
        assert workbook["metadata"]["B5"].value == "MAD"
        assert workbook["risk_register"]["A2"].value == "Études"
        assert workbook["risk_register"]["B2"].value == "triangular"
        assert workbook["risk_register"]["C2"].value == 10_000
        assert workbook["risk_register"].auto_filter.ref == "A1:N2"
        assert len(workbook["risk_register"].data_validations.dataValidation) == 2
    finally:
        workbook.close()
