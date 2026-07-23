import importlib.util
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from monte_carlo_simulator.analysis.sensitivity import build_tornado_data
from monte_carlo_simulator.analysis.value_at_risk import compute_value_at_risk
from monte_carlo_simulator.application import run_simulation_from_excel
from monte_carlo_simulator.distributions import (
    EventDistribution,
    LogNormalDistribution,
    NormalDistribution,
    PertDistribution,
    TriangularDistribution,
    UniformDistribution,
)
from monte_carlo_simulator.engine import GaussianCopulaSampler
from monte_carlo_simulator.engine.aggregation import aggregate_total
from monte_carlo_simulator.engine.convergence import check_convergence
from monte_carlo_simulator.engine.correlations import apply_correlation_matrix
from monte_carlo_simulator.engine.simulator import MonteCarloSimulator
from monte_carlo_simulator.exceptions import RiskRegisterValidationError, ValidationError
from monte_carlo_simulator.io import load_risk_register
from monte_carlo_simulator.models import CorrelationMatrix, RiskItem, SimulationConfig
from monte_carlo_simulator.visualization.s_curve import build_s_curve
from monte_carlo_simulator.visualization.tornado import build_tornado_chart

_spec = importlib.util.spec_from_file_location(
    "generate_correlated_cost_register",
    Path(__file__).parents[2] / "scripts" / "generate_correlated_cost_register.py",
)
assert _spec is not None and _spec.loader is not None
_module = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_module)
generate = _module.generate


def _add_correlations(path: Path, headers: list[object], rows: list[list[object]]) -> None:
    workbook = load_workbook(path)
    sheet = workbook.create_sheet("correlations")
    sheet.append([None, *headers])
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def test_excel_correlation_axes_are_aligned_independently(excel_workbook_factory) -> None:
    path = excel_workbook_factory(
        rows=[
            {"name": "A", "distribution": "uniform", "minimum": 0, "maximum": 1},
            {"name": "B", "distribution": "uniform", "minimum": 0, "maximum": 1},
            {"name": "C", "distribution": "uniform", "minimum": 0, "maximum": 1},
        ]
    )
    semantic = {
        ("A", "A"): 1,
        ("B", "B"): 1,
        ("C", "C"): 1,
        ("A", "B"): 0.2,
        ("B", "A"): 0.2,
        ("A", "C"): -0.1,
        ("C", "A"): -0.1,
        ("B", "C"): 0.3,
        ("C", "B"): 0.3,
    }
    columns = ["C", "A", "B"]
    rows = [[name, *[semantic[(name, column)] for column in columns]] for name in ["B", "C", "A"]]
    _add_correlations(path, columns, rows)

    register = load_risk_register(path)

    assert register.correlation_matrix is not None
    np.testing.assert_allclose(
        register.correlation_matrix.values,
        np.array([[1, 0.2, -0.1], [0.2, 1, 0.3], [-0.1, 0.3, 1]]),
    )


@pytest.mark.parametrize("bad", [True, "0.5", 1 + 2j, np.nan, np.inf, -0.1, 1.1])
def test_ppf_rejects_invalid_probabilities(bad: object) -> None:
    with pytest.raises(ValidationError):
        UniformDistribution(0, 10).ppf(np.array([bad]))


@pytest.mark.parametrize(
    "distribution",
    [
        UniformDistribution(0, 10),
        NormalDistribution(5, 1),
    ],
)
def test_ppf_preserves_shape_and_boundaries(distribution) -> None:
    probabilities = np.array([[0.0, 0.5], [0.75, 1.0]])
    values = distribution.ppf(probabilities)
    assert values.shape == probabilities.shape


def test_correlation_matrix_and_sampler_validation() -> None:
    matrix = CorrelationMatrix(("A", "B"), [[1, -0.4], [-0.4, 1]])
    sampler = GaussianCopulaSampler(matrix)
    rng = np.random.default_rng(123)
    samples = sampler.sample([NormalDistribution(0, 1), NormalDistribution(0, 1)], rng, 2000)
    assert samples.shape == (2000, 2)
    assert np.corrcoef(samples.T)[0, 1] < -0.25
    with pytest.raises(ValidationError):
        sampler.sample([NormalDistribution(0, 1)], rng, 10)
    with pytest.raises(ValidationError):
        sampler.sample([NormalDistribution(0, 1), NormalDistribution(0, 1)], rng, 0)
    with pytest.raises(ValidationError):
        sampler.sample([NormalDistribution(0, 1), NormalDistribution(0, 1)], object(), 10)
    with pytest.raises(ValidationError):
        CorrelationMatrix(("A", "B"), [[1, 2], [2, 1]])


def test_invalid_excel_correlation_context(excel_workbook_factory) -> None:
    path = excel_workbook_factory(
        rows=[
            {"name": "A", "distribution": "uniform", "minimum": 0, "maximum": 1},
            {"name": "B", "distribution": "uniform", "minimum": 0, "maximum": 1},
        ]
    )
    _add_correlations(path, ["A", "A"], [["A", 1, 0.2], ["B", None, 1]])
    with pytest.raises(RiskRegisterValidationError) as caught:
        load_risk_register(path)
    assert {issue.sheet for issue in caught.value.issues} == {"correlations"}
    assert any(issue.row == 1 and issue.value == "A" for issue in caught.value.issues)
    assert any(issue.row == 3 and issue.value is None for issue in caught.value.issues)


def test_synthetic_correlated_example_regenerates_and_runs(tmp_path: Path) -> None:
    path = generate(tmp_path / "correlated_cost_register.xlsx")
    run = run_simulation_from_excel(
        path,
        SimulationConfig(number_of_simulations=10_000, random_seed=7),
        tmp_path / "output",
    )
    assert len(run.result.samples) == 10_000
    assert run.summary_path.exists()
    assert run.histogram_path.exists()


@pytest.mark.parametrize(
    "distribution",
    [
        TriangularDistribution(0, 5, 10),
        PertDistribution(0, 5, 10),
        UniformDistribution(0, 10),
        NormalDistribution(5, 1),
        LogNormalDistribution(5, 1),
        EventDistribution(0.5, 10),
    ],
)
def test_all_distribution_ppfs_accept_1d_and_2d_probabilities(distribution) -> None:
    one_dimensional = np.array([0.0, 0.25, 0.5, 0.75, 1.0])
    two_dimensional = np.array([[0.0, 0.5], [0.75, 1.0]])
    assert distribution.ppf(one_dimensional).shape == one_dimensional.shape
    assert distribution.ppf(two_dimensional).shape == two_dimensional.shape


@pytest.mark.parametrize(
    "values",
    [
        np.array([1, 0, 0], dtype=bool),
        np.array(["0.1"]),
        np.array([1 + 0j]),
        np.array([np.nan]),
        np.array([np.inf]),
    ],
)
def test_all_distribution_ppfs_reject_invalid_probability_arrays(values: np.ndarray) -> None:
    distributions = [
        TriangularDistribution(0, 5, 10),
        PertDistribution(0, 5, 10),
        UniformDistribution(0, 10),
        NormalDistribution(5, 1),
        LogNormalDistribution(5, 1),
        EventDistribution(0.5, 10),
    ]
    for distribution in distributions:
        with pytest.raises(ValidationError):
            distribution.ppf(values)


@pytest.mark.parametrize(
    "names, values, message",
    [
        (("A",), [1.0], "two-dimensional"),
        (("A", "B"), [[1.0, 0.2, 0.1], [0.2, 1.0, 0.1]], "square"),
        (("A", "B", "C"), [[1, 0], [0, 1]], "dimension"),
        (("A", "B"), [[1, True], [True, 1]], "finite real"),
        (("A", "B"), [[1, "0.2"], ["0.2", 1]], "finite real"),
        (("A", "B"), [[1, 1 + 0j], [1 + 0j, 1]], "finite real"),
        (("A", "B"), [[1, np.nan], [np.nan, 1]], "finite real"),
        (("A", "B"), [[1, np.inf], [np.inf, 1]], "finite real"),
        (("A", "B"), [[1, 1.2], [1.2, 1]], "[-1, 1]"),
        (("A", "B"), [[0.999, 0], [0, 1]], "diagonal"),
        (("A", "B"), [[1, 0.1], [0.2, 1]], "symmetric"),
        (("A", "B"), [[1, 1], [1, 1]], "positive definite"),
        (("A", "B"), [[1, -1], [-1, 1]], "positive definite"),
        (("A", "B", "C"), [[1, 0.9, 0.9], [0.9, 1, -0.9], [0.9, -0.9, 1]], "positive definite"),
    ],
)
def test_correlation_matrix_rejects_invalid_values(names, values, message: str) -> None:
    with pytest.raises(ValidationError, match=message):
        CorrelationMatrix(names, values)


@pytest.mark.parametrize("names", [("",), ("  ",), (1,), ("A", " a ")])
def test_correlation_matrix_rejects_invalid_names(names) -> None:
    with pytest.raises(ValidationError):
        CorrelationMatrix(names, [[1]])


def test_correlation_matrix_copies_and_freezes_values() -> None:
    values = np.array([[1.0, 0.2], [0.2, 1.0]])
    matrix = CorrelationMatrix(("A", "B"), values)
    values[0, 1] = 0.8
    assert matrix.values[0, 1] == 0.2
    with pytest.raises(ValueError):
        matrix.values[0, 1] = 0.3


def test_correlation_matrix_aligned_to_reorders_by_name() -> None:
    matrix = CorrelationMatrix(("B", "A", "C"), [[1, 0.2, -0.1], [0.2, 1, 0.3], [-0.1, 0.3, 1]])
    aligned = matrix.aligned_to(("A", "B", "C"))
    assert aligned.item_names == ("A", "B", "C")
    np.testing.assert_allclose(aligned.values, [[1, 0.2, 0.3], [0.2, 1, -0.1], [0.3, -0.1, 1]])
    with pytest.raises(ValidationError, match="match exactly"):
        matrix.aligned_to(("A", "B", "D"))


def test_sampler_supports_1x1_identity_3x3_and_apply_wrapper() -> None:
    rng = np.random.default_rng(321)
    one = GaussianCopulaSampler(CorrelationMatrix(("A",), [[1]])).sample(
        [NormalDistribution(0, 1)], rng, 50
    )
    assert one.shape == (50, 1)
    identity = CorrelationMatrix(("A", "B"), [[1, 0], [0, 1]])
    wrapped = apply_correlation_matrix(
        [NormalDistribution(0, 1), NormalDistribution(0, 1)], identity, rng, 50
    )
    assert wrapped.shape == (50, 2)
    three = CorrelationMatrix(("A", "B", "C"), [[1, 0.3, -0.2], [0.3, 1, 0.1], [-0.2, 0.1, 1]])
    samples = GaussianCopulaSampler(three).sample(
        [NormalDistribution(0, 1), NormalDistribution(0, 1), NormalDistribution(0, 1)],
        rng,
        4_000,
    )
    correlations = np.corrcoef(samples.T)
    assert correlations[0, 1] > 0.2
    assert correlations[0, 2] < -0.1


class ExtremeRng:
    def standard_normal(self, size):
        return np.full(size, 40.0)


def test_copula_extreme_probabilities_are_clipped_to_finite_ppf_values() -> None:
    matrix = CorrelationMatrix(("A",), [[1]])
    samples = GaussianCopulaSampler(matrix).sample([NormalDistribution(0, 1)], ExtremeRng(), 5)
    assert np.all(np.isfinite(samples))


def test_monte_carlo_simulator_correlated_api_and_traceability(sample_config) -> None:
    items = [
        RiskItem("A", "normal", mean=0, standard_deviation=1),
        RiskItem("B", "normal", mean=0, standard_deviation=1),
        RiskItem("C", "normal", mean=0, standard_deviation=1),
    ]
    matrix = CorrelationMatrix(("C", "A", "B"), [[1, -0.2, 0.1], [-0.2, 1, 0.4], [0.1, 0.4, 1]])
    result = MonteCarloSimulator().run(items, sample_config, correlation_matrix=matrix)
    assert list(result.item_samples.columns) == ["A", "B", "C"]
    assert result.correlation_item_names == ("A", "B", "C")
    np.testing.assert_allclose(
        result.correlation_matrix,
        [[1, 0.4, -0.2], [0.4, 1, 0.1], [-0.2, 0.1, 1]],
    )
    result.correlation_matrix[0, 1] = 0.0
    assert matrix.values[1, 2] == 0.4


def test_monte_carlo_without_correlation_keeps_historical_rng_sequence(
    sample_items, sample_config
) -> None:
    simulator = MonteCarloSimulator()
    historical = simulator.run(sample_items, sample_config)
    explicit_none = simulator.run(sample_items, sample_config, correlation_matrix=None)
    np.testing.assert_allclose(historical.samples, explicit_none.samples)
    assert explicit_none.correlation_item_names is None
    assert explicit_none.correlation_matrix is None


def test_deterministic_item_allows_identity_but_rejects_non_zero_correlation(sample_config) -> None:
    items = [
        RiskItem("A", "uniform", minimum=5, maximum=5),
        RiskItem("B", "normal", mean=0, standard_deviation=1),
    ]
    identity = CorrelationMatrix(("A", "B"), [[1, 0], [0, 1]])
    result = MonteCarloSimulator().run(items, sample_config, correlation_matrix=identity)
    assert np.all(result.item_samples["A"] == 5)
    non_zero = CorrelationMatrix(("A", "B"), [[1, 0.2], [0.2, 1]])
    with pytest.raises(ValidationError, match="Deterministic"):
        MonteCarloSimulator().run(items, sample_config, correlation_matrix=non_zero)


def test_workflow_without_matrix_has_no_traceability(
    excel_workbook_factory, tmp_path: Path
) -> None:
    path = excel_workbook_factory()
    run = run_simulation_from_excel(path, SimulationConfig(number_of_simulations=50), tmp_path)
    assert run.result.correlation_item_names is None
    assert run.result.correlation_matrix is None


def test_small_utility_and_placeholder_modules_are_exercised() -> None:
    frame = np.array([1.0, 2.0, 3.0])
    assert compute_value_at_risk(frame, 0.5) == 2.0
    item_samples = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    assert aggregate_total(item_samples).tolist() == [4, 6]
    for function in [build_tornado_data, check_convergence, build_s_curve, build_tornado_chart]:
        with pytest.raises(NotImplementedError):
            function()
