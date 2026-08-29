"""Contract tests for the React-to-Python HTTP adapter."""

from collections.abc import Iterator
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from monte_carlo_simulator.io import create_risk_register_workbook
from monte_carlo_simulator.storage import database
from monte_carlo_simulator.web_api import app


@pytest.fixture(autouse=True)
def isolated_database(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv(database.DATA_DIR_ENV, str(tmp_path / "data"))


@pytest.fixture
def client() -> Iterator[TestClient]:
    with TestClient(app) as http:
        yield http


def _draft_payload() -> dict[str, object]:
    return {
        "schemaVersion": "1.0",
        "metadata": {
            "projectName": "Projet interface",
            "analysisType": "cost",
            "defaultUnit": "EUR",
            "baselineEstimate": 1000,
            "description": "Brouillon JSON",
        },
        "items": [
            {
                "id": "risk-1",
                "name": "Études",
                "distribution": "triangular",
                "minimum": 100,
                "mostLikely": 150,
                "maximum": 250,
                "category": "Ingénierie",
                "unit": "EUR",
                "enabled": True,
                "notes": "",
            },
            {
                "id": "risk-2",
                "name": "Aléa",
                "distribution": "event",
                "probability": 0.2,
                "impact": 500,
                "category": "Marché",
                "unit": "EUR",
                "enabled": True,
                "notes": "",
            },
        ],
        "correlations": {
            "mode": "correlated",
            "names": ["Études", "Aléa"],
            "values": [[1, 0.2], [0.2, 1]],
        },
    }


def test_health_reports_a_working_engine(client: TestClient) -> None:
    response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "engine": "monte-carlo-simulator"}


def test_shutdown_requires_the_packaged_launcher(client: TestClient) -> None:
    response = client.post("/api/shutdown")

    assert response.status_code == 409
    assert "application portable" in response.json()["detail"]


def test_simulation_rejects_non_xlsx_upload(client: TestClient) -> None:
    response = client.post(
        "/api/simulate",
        files={"file": ("register.csv", b"name,value", "text/csv")},
    )

    assert response.status_code == 415
    assert response.json()["detail"] == "Seuls les registres .xlsx sont acceptés."


def test_template_endpoint_returns_excel_workbook(client: TestClient) -> None:
    response = client.get("/api/template")

    assert response.status_code == 200
    assert response.headers["content-disposition"].endswith(
        'filename="risk_register_template.xlsx"'
    )
    assert response.content.startswith(b"PK")


def test_simulation_endpoint_runs_engine_and_preserves_correlations(
    client: TestClient, tmp_path: Path
) -> None:
    workbook_path = tmp_path / "correlated.xlsx"
    create_risk_register_workbook(
        workbook_path,
        metadata_values={
            "project_name": "API correlated",
            "analysis_type": "cost",
            "default_unit": "EUR",
            "baseline_estimate": 120,
        },
        risk_rows=[
            {"name": "A", "distribution": "uniform", "minimum": 0, "maximum": 100},
            {"name": "B", "distribution": "uniform", "minimum": 0, "maximum": 200},
        ],
    )
    workbook = load_workbook(workbook_path)
    sheet = workbook.create_sheet("correlations")
    sheet.append([None, "A", "B"])
    sheet.append(["A", 1, 0.25])
    sheet.append(["B", 0.25, 1])
    workbook.save(workbook_path)
    workbook.close()

    with workbook_path.open("rb") as stream:
        response = client.post(
            "/api/simulate",
            files={
                "file": (
                    workbook_path.name,
                    stream,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={
                "simulations": "200",
                "seed": "20260818",
                "confidence_levels": "0.50,0.80,0.90,0.95",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["project"]["name"] == "API correlated"
    assert payload["run"]["simulations"] == 200
    assert payload["run"]["correlationsEnabled"] is True
    assert [row["percentile"] for row in payload["percentiles"]] == [
        "P50",
        "P80",
        "P90",
        "P95",
    ]
    assert payload["histogram"]
    assert payload["sCurve"]


def test_register_draft_can_be_validated_exported_and_imported(client: TestClient) -> None:
    draft = _draft_payload()

    validation = client.post("/api/register/validate", json=draft)
    assert validation.status_code == 200
    assert validation.json() == {
        "valid": True,
        "projectName": "Projet interface",
        "totalItems": 2,
        "activeItems": 2,
        "correlationsEnabled": True,
    }

    exported = client.post("/api/register/export", json=draft)
    assert exported.status_code == 200
    assert exported.content.startswith(b"PK")

    imported = client.post(
        "/api/register/import",
        files={
            "file": (
                "registre.xlsx",
                exported.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200
    imported_register = imported.json()["register"]
    assert imported_register["metadata"]["projectName"] == "Projet interface"
    assert [item["name"] for item in imported_register["items"]] == ["Études", "Aléa"]
    assert imported_register["correlations"]["values"] == [[1.0, 0.2], [0.2, 1.0]]


def test_register_draft_can_launch_simulation(client: TestClient) -> None:
    response = client.post(
        "/api/register/simulate",
        json={
            "register": _draft_payload(),
            "config": {
                "simulations": 120,
                "seed": 20260820,
                "levels": [50, 80, 90, 95],
                "decisionPercentile": 80,
                "exceedanceThreshold": 1800,
                "convergenceTolerance": 1,
            },
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["project"]["name"] == "Projet interface"
    assert payload["run"]["simulations"] == 120
    assert payload["run"]["correlationsEnabled"] is True


def test_complete_results_can_be_exported_to_excel(client: TestClient) -> None:
    config = {
        "simulations": 120,
        "seed": 20260820,
        "levels": [50, 80, 90, 95],
        "decisionPercentile": 80,
        "exceedanceThreshold": 1800,
        "convergenceTolerance": 1,
    }
    simulated = client.post(
        "/api/register/simulate",
        json={"register": _draft_payload(), "config": config},
    )
    assert simulated.status_code == 200

    exported = client.post(
        "/api/results/export",
        json={
            "result": simulated.json(),
            "register": _draft_payload(),
            "config": config,
        },
    )

    assert exported.status_code == 200
    assert exported.content.startswith(b"PK")
    assert exported.headers["content-disposition"].endswith('filename="resultats_monte_carlo.xlsx"')
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    assert workbook.sheetnames == [
        "Synthèse",
        "Percentiles",
        "Sensibilité",
        "Convergence",
        "Hypothèses",
        "Robustesse",
    ]
    assert workbook["Synthèse"]["B5"].value == "Projet interface"
    assert workbook["Hypothèses"]["B5"].value == 120
    assert workbook["Robustesse"]["A14"].value == "Études"
    assert workbook["Robustesse"]["B14"].value == 1
    workbook.close()

    bundle = client.post(
        "/api/results/export-bundle",
        json={
            "result": simulated.json(),
            "register": _draft_payload(),
            "config": config,
        },
    )
    assert bundle.status_code == 200
    assert bundle.headers["content-type"] == "application/zip"
    with ZipFile(BytesIO(bundle.content)) as archive:
        expected = {
            "LISEZ_MOI.txt",
            "registre_risques_utilise.xlsx",
            "resultats_monte_carlo.xlsx",
            "donnees/resultats.json",
            "graphiques/01_histogramme_distribution.png",
            "graphiques/02_courbe_probabilite_s.png",
            "graphiques/03_sensibilite_tornado.png",
            "graphiques/04_convergence.png",
        }
        assert set(archive.namelist()) == expected
        assert archive.read("registre_risques_utilise.xlsx").startswith(b"PK")
        assert archive.read("resultats_monte_carlo.xlsx").startswith(b"PK")
        for name in expected:
            if name.endswith(".png"):
                assert archive.read(name).startswith(b"\x89PNG\r\n\x1a\n")
