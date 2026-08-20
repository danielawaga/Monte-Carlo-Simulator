"""Tests for the Excel-to-editable-model boundary used by the UI."""

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from monte_carlo_simulator.application.hypotheses import load_editable_risk_register
from monte_carlo_simulator.io import create_risk_register_workbook


def test_excel_register_becomes_editable_rows_and_keeps_disabled_items(tmp_path: Path) -> None:
    path = tmp_path / "editable.xlsx"
    create_risk_register_workbook(
        path,
        metadata_values={
            "project_name": "Scénario importé",
            "analysis_type": "cost",
            "default_unit": "EUR",
            "baseline_estimate": 1_000,
        },
        risk_rows=[
            {
                "name": "Actif",
                "distribution": "triangular",
                "minimum": 10,
                "most_likely": 20,
                "maximum": 40,
                "enabled": True,
            },
            {
                "name": "Option désactivée",
                "distribution": "event",
                "probability": 0.2,
                "impact": 100,
                "enabled": False,
            },
        ],
    )

    editable = load_editable_risk_register(path)

    assert editable.metadata.project_name == "Scénario importé"
    assert editable.metadata.baseline_estimate == 1_000
    assert editable.rows["name"].tolist() == ["Actif", "Option désactivée"]
    assert editable.rows["enabled"].tolist() == [True, False]
    assert "_excel_row" not in editable.rows


def test_editable_register_preserves_validated_correlation_matrix(tmp_path: Path) -> None:
    path = tmp_path / "editable-correlated.xlsx"
    create_risk_register_workbook(
        path,
        metadata_values={
            "project_name": "Scénario corrélé",
            "analysis_type": "cost",
            "default_unit": "EUR",
            "baseline_estimate": 1_000,
        },
        risk_rows=[
            {"name": "A", "distribution": "uniform", "minimum": 0, "maximum": 100},
            {"name": "B", "distribution": "uniform", "minimum": 0, "maximum": 200},
        ],
    )
    workbook = load_workbook(path)
    sheet = workbook.create_sheet("correlations")
    sheet.append([None, "A", "B"])
    sheet.append(["A", 1, 0.35])
    sheet.append(["B", 0.35, 1])
    workbook.save(path)
    workbook.close()

    editable = load_editable_risk_register(path)

    assert editable.correlation_matrix is not None
    assert editable.correlation_matrix.item_names == ("A", "B")
    np.testing.assert_allclose(editable.correlation_matrix.values, [[1, 0.35], [0.35, 1]])
