import os
from collections.abc import Callable, Mapping, Sequence
from pathlib import Path

import pytest
from openpyxl import Workbook

from monte_carlo_simulator.io.schema import METADATA_KEYS, RISK_REGISTER_COLUMNS
from monte_carlo_simulator.models import RiskItem, SimulationConfig

os.environ.setdefault("PYTHONPATH", "src")

ExcelWorkbookFactory = Callable[..., Path]

_DEFAULT_METADATA: dict[str, object] = {
    "schema_version": "1.0",
    "project_name": "Test Project",
    "analysis_type": "cost",
    "default_unit": "EUR",
    "baseline_estimate": 1_000,
    "description": "Fictitious test data.",
}

_DEFAULT_RISK_ROW: dict[str, object] = {
    "name": "Uniform item",
    "distribution": "uniform",
    "minimum": 10,
    "maximum": 20,
    "category": "Test",
    "unit": None,
    "enabled": True,
    "notes": "Synthetic test row.",
}


@pytest.fixture
def sample_items() -> list[RiskItem]:
    return [
        RiskItem("A", "triangular", minimum=1.0, most_likely=2.0, maximum=4.0),
        RiskItem("B", "triangular", minimum=2.0, most_likely=3.0, maximum=5.0),
        RiskItem("C", "triangular", minimum=3.0, most_likely=4.0, maximum=7.0),
        RiskItem("D", "triangular", minimum=1.0, most_likely=2.0, maximum=3.0),
        RiskItem("E", "triangular", minimum=2.0, most_likely=2.5, maximum=3.0),
    ]


@pytest.fixture
def sample_config() -> SimulationConfig:
    return SimulationConfig(number_of_simulations=10_000, random_seed=42)


@pytest.fixture
def excel_workbook_factory(tmp_path: Path) -> ExcelWorkbookFactory:
    """Build schema-v1 workbooks with openpyxl for isolated IO tests."""

    def factory(
        filename: str = "risk_register.xlsx",
        *,
        metadata: Mapping[str, object] | None = None,
        rows: Sequence[Mapping[str, object]] | None = None,
        include_metadata: bool = True,
        include_risk_register: bool = True,
        omitted_columns: Sequence[str] = (),
        omitted_metadata_keys: Sequence[str] = (),
    ) -> Path:
        path = tmp_path / filename
        workbook = Workbook()
        workbook.remove(workbook.active)

        if include_metadata:
            metadata_sheet = workbook.create_sheet("metadata")
            metadata_sheet.append(["key", "value"])
            metadata_values = {**_DEFAULT_METADATA, **(metadata or {})}
            for key in METADATA_KEYS:
                if key not in omitted_metadata_keys:
                    metadata_sheet.append([key, metadata_values[key]])

        if include_risk_register:
            register_sheet = workbook.create_sheet("risk_register")
            columns = [column for column in RISK_REGISTER_COLUMNS if column not in omitted_columns]
            register_sheet.append(columns)
            risk_rows = [_DEFAULT_RISK_ROW] if rows is None else rows
            for risk_row in risk_rows:
                register_sheet.append([risk_row.get(column) for column in columns])

        if not workbook.sheetnames:
            workbook.create_sheet("placeholder")
        workbook.save(path)
        workbook.close()
        return path

    return factory
