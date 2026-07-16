"""Generate the public schema-v1 Excel risk-register template."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from monte_carlo_simulator.io.schema import (
    INSTRUCTIONS_SHEET,
    METADATA_KEYS,
    METADATA_SHEET,
    RISK_REGISTER_COLUMNS,
    RISK_REGISTER_SHEET,
    SCHEMA_VERSION,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_DISABLED_FILL = PatternFill("solid", fgColor="E7E6E6")


def create_risk_register_template(output_path: str | Path) -> Path:
    """Create a documented, validated example workbook for schema version 1.0."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    metadata = workbook.active
    metadata.title = METADATA_SHEET
    risk_register = workbook.create_sheet(RISK_REGISTER_SHEET)
    instructions = workbook.create_sheet(INSTRUCTIONS_SHEET)

    _populate_metadata(metadata)
    _populate_risk_register(risk_register)
    _populate_instructions(instructions)

    workbook.properties.title = "Monte Carlo Simulator risk register template"
    workbook.properties.subject = f"Excel risk-register schema {SCHEMA_VERSION}"
    workbook.properties.description = "Fictitious example data only; no confidential content."
    workbook.save(path)
    workbook.close()
    return path


def _populate_metadata(worksheet: Worksheet) -> None:
    worksheet.append(["key", "value", "description"])
    values = {
        "schema_version": SCHEMA_VERSION,
        "project_name": "Fictitious Infrastructure Project",
        "analysis_type": "cost",
        "default_unit": "EUR",
        "baseline_estimate": 250_000,
        "description": "Public example containing fictitious values only.",
    }
    descriptions = {
        "schema_version": "Required. Must remain 1.0 for this template.",
        "project_name": "Required project or analysis name.",
        "analysis_type": "Required: cost or duration.",
        "default_unit": "Required common unit inherited by blank row units.",
        "baseline_estimate": "Optional finite reference value; not added to simulated totals.",
        "description": "Optional non-confidential context.",
    }
    for key in METADATA_KEYS:
        worksheet.append([key, values[key], descriptions[key]])

    _style_header(worksheet, 3)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 38
    worksheet.column_dimensions["C"].width = 68
    worksheet["B6"].number_format = "#,##0.00"

    analysis_validation = DataValidation(
        type="list",
        formula1='"cost,duration"',
        allow_blank=False,
    )
    analysis_validation.error = "Choose cost or duration."
    analysis_validation.errorTitle = "Invalid analysis type"
    analysis_validation.prompt = "Select the type of project analysis."
    analysis_validation.promptTitle = "Analysis type"
    analysis_validation.showErrorMessage = True
    analysis_validation.showInputMessage = True
    worksheet.add_data_validation(analysis_validation)
    analysis_validation.add(worksheet["B4"])


def _populate_risk_register(worksheet: Worksheet) -> None:
    worksheet.append(list(RISK_REGISTER_COLUMNS))
    rows = [
        [
            "Preliminary studies",
            "triangular",
            10_000,
            15_000,
            22_000,
            None,
            None,
            None,
            None,
            None,
            "Engineering",
            "EUR",
            True,
            "Three-point estimate.",
        ],
        [
            "Detailed design",
            "pert",
            20_000,
            30_000,
            50_000,
            None,
            None,
            None,
            None,
            4,
            "Engineering",
            None,
            True,
            "Blank unit inherits EUR; lambda defaults to 4 when omitted.",
        ],
        [
            "Permitting",
            "uniform",
            5_000,
            None,
            12_000,
            None,
            None,
            None,
            None,
            None,
            "Regulatory",
            None,
            True,
            "All values between minimum and maximum are equally likely.",
        ],
        [
            "Site supervision",
            "normal",
            None,
            None,
            None,
            18_000,
            3_000,
            None,
            None,
            None,
            "Delivery",
            None,
            True,
            "Arithmetic mean and standard deviation.",
        ],
        [
            "Commodity escalation",
            "lognormal",
            None,
            None,
            None,
            25_000,
            7_500,
            None,
            None,
            None,
            "Market",
            None,
            True,
            "Mean and standard deviation are in arithmetic space.",
        ],
        [
            "Regulatory change",
            "event",
            None,
            None,
            None,
            None,
            None,
            0.20,
            40_000,
            None,
            "Regulatory",
            None,
            True,
            "Bernoulli occurrence multiplied by a deterministic impact.",
        ],
    ]
    for row in rows:
        worksheet.append(row)

    _style_header(worksheet, len(RISK_REGISTER_COLUMNS))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:N{worksheet.max_row}"
    worksheet.row_dimensions[1].height = 30

    widths = (28, 16, 14, 16, 14, 14, 20, 14, 14, 14, 18, 12, 12, 62)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 14)
        for column in range(3, 11):
            row[column - 1].number_format = "#,##0.00"
    for cell in worksheet["H"][1:]:
        cell.number_format = "0.00%"

    distribution_validation = DataValidation(
        type="list",
        formula1='"triangular,pert,uniform,normal,lognormal,event"',
        allow_blank=False,
    )
    distribution_validation.error = "Choose one of the six supported distributions."
    distribution_validation.errorTitle = "Invalid distribution"
    distribution_validation.showErrorMessage = True
    worksheet.add_data_validation(distribution_validation)
    distribution_validation.add("B2:B1000")

    enabled_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True,
    )
    enabled_validation.error = "Choose TRUE or FALSE, or leave the cell blank for TRUE."
    enabled_validation.errorTitle = "Invalid enabled value"
    enabled_validation.showErrorMessage = True
    worksheet.add_data_validation(enabled_validation)
    enabled_validation.add("M2:M1000")

    disabled_rule = FormulaRule(formula=["$M2=FALSE"], fill=_DISABLED_FILL)
    worksheet.conditional_formatting.add("A2:N1000", disabled_rule)


def _populate_instructions(worksheet: Worksheet) -> None:
    rows = [
        ("Risk-register schema 1.0", "How to use this public, non-confidential template."),
        (
            "Workflow",
            "Complete metadata and one active risk per row, save as .xlsx, then run the CLI.",
        ),
        (
            "Confidentiality",
            "Never commit real client risk registers. Use anonymized fictitious data only.",
        ),
        (
            "Blank cells",
            "Whitespace-only cells and empty cells are treated as blank. "
            "Unused parameters stay blank.",
        ),
        ("enabled", "Blank or TRUE means active; FALSE ignores the entire row."),
        (
            "unit",
            "A blank row unit inherits default_unit. "
            "Every active item must resolve to the same unit.",
        ),
        (
            "baseline_estimate",
            "Optional reference estimate. It is returned as metadata and is never "
            "added automatically.",
        ),
        (
            "triangular",
            "Required: minimum, most_likely, maximum with minimum <= most_likely <= maximum.",
        ),
        (
            "pert",
            "Required: minimum, most_likely, maximum. lambda_shape is optional, "
            "defaults to 4 and must be > 0.",
        ),
        ("uniform", "Required: minimum and maximum with minimum <= maximum."),
        ("normal", "Required: arithmetic mean and standard_deviation >= 0."),
        ("lognormal", "Required: arithmetic mean > 0 and arithmetic standard_deviation >= 0."),
        (
            "event",
            "Required: probability in [0, 1] and deterministic impact; "
            "negative impact can model an opportunity.",
        ),
        (
            "Numeric text",
            "Finite numeric strings are accepted; booleans, NaN, infinities and "
            "arbitrary text are rejected.",
        ),
        (
            "CLI",
            "python -m monte_carlo_simulator.cli --input PATH --simulations 10000 "
            "--seed 42 --output-dir data/output",
        ),
    ]
    for instruction_row in rows:
        worksheet.append(instruction_row)
    _style_header(worksheet, 2)
    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row_number, 1).font = Font(bold=True)
        worksheet.cell(row_number, 1).fill = _SECTION_FILL
        worksheet.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 105
    worksheet.freeze_panes = "A2"


def _style_header(worksheet: Worksheet, column_count: int) -> None:
    for cell in worksheet[1][:column_count]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
