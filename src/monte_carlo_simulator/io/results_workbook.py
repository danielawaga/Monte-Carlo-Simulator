"""Excel export for a complete web simulation result."""

from __future__ import annotations

from collections.abc import Iterable, Mapping
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ORANGE = "F45116"
NAVY = "14213D"
MUTED = "69758D"
LIGHT_BLUE = "EAF2FC"
LIGHT_ORANGE = "FFF1E8"
LIGHT_GRAY = "F1F4F8"
WHITE = "FFFFFF"
BORDER = "D9E0E8"

THIN_BORDER = Border(bottom=Side(style="thin", color=BORDER))

LABELS = {
    "mean": "Moyenne",
    "std": "Écart-type",
    "minimum": "Minimum observé",
    "maximum": "Maximum observé",
    "percentile": "Percentile",
    "amount": "Montant / durée",
    "recommended_reserve": "Réserve recommandée",
    "item_name": "Poste",
    "spearman_rho": "Coefficient de Spearman",
    "draw_count": "Nombre de tirages",
    "target_percentile": "Percentile contrôlé",
    "relative_change": "Variation relative",
    "stop_recommended": "Arrêt recommandé",
    "baseline": "Valeur de référence",
    "simulated_mean": "Moyenne simulée",
    "mean_difference": "Écart à la moyenne",
    "exceedance_probability": "Probabilité de dépassement",
    "minimum_eigenvalue": "Valeur propre minimale",
    "condition_number": "Conditionnement",
    "automatic_repair_applied": "Réparation automatique",
}


def _display(value: Any) -> Any:
    if value is None:
        return "Non disponible"
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (str, int, float)):
        return value
    return str(value)


def _title(sheet: Worksheet, title: str, subtitle: str, width: int) -> int:
    last_column = get_column_letter(max(width, 2))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 31
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, italic=True, color=MUTED)
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 27
    sheet.sheet_view.showGridLines = False
    return 4


def _section(sheet: Worksheet, row: int, label: str, width: int) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(width, 2))
    cell = sheet.cell(row=row, column=1, value=label)
    cell.font = Font(size=11, bold=True, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 23
    return row + 1


def _key_values(sheet: Worksheet, row: int, values: Iterable[tuple[str, Any]]) -> int:
    for label, value in values:
        label_cell = sheet.cell(row=row, column=1, value=label)
        value_cell = sheet.cell(row=row, column=2, value=_display(value))
        label_cell.font = Font(bold=True, color=NAVY)
        value_cell.font = Font(color=NAVY)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        label_cell.border = value_cell.border = THIN_BORDER
        label_cell.alignment = value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row + 1


def _record_table(
    sheet: Worksheet,
    row: int,
    records: list[Mapping[str, Any]],
    *,
    empty_message: str = "Aucune donnée disponible",
) -> int:
    if not records:
        sheet.cell(row=row, column=1, value=empty_message).font = Font(italic=True, color=MUTED)
        return row + 2

    keys = list(dict.fromkeys(key for record in records for key in record))
    for column, key in enumerate(keys, start=1):
        cell = sheet.cell(
            row=row, column=column, value=LABELS.get(key, key.replace("_", " ").title())
        )
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    header_row = row
    row += 1
    for record in records:
        for column, key in enumerate(keys, start=1):
            cell = sheet.cell(row=row, column=column, value=_display(record.get(key)))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(record.get(key), float):
                cell.number_format = "#,##0.000"
        row += 1
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(keys))}{row - 1}"
    return row + 1


def _finish(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A4"
    for index, column in enumerate(sheet.columns, start=1):
        values = [str(cell.value) for cell in column if cell.value is not None]
        width = max((len(value) for value in values), default=10) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 12), 34)


def _records(value: Any) -> list[Mapping[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, Mapping)]


def _mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def build_simulation_results_workbook(
    *, result: Mapping[str, Any], register: Mapping[str, Any], config: Mapping[str, Any]
) -> bytes:
    """Return a styled, auditable XLSX workbook for one simulation run."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    project = _mapping(result.get("project"))
    run = _mapping(result.get("run"))
    metadata = _mapping(register.get("metadata"))

    summary = workbook.create_sheet("Synthèse")
    row = _title(
        summary,
        "Résultats de la simulation Monte-Carlo",
        "Synthèse décisionnelle et traçabilité de l'exécution",
        2,
    )
    row = _section(summary, row, "Projet et exécution", 2)
    row = _key_values(
        summary,
        row,
        [
            ("Projet", project.get("name")),
            ("Type d'analyse", project.get("analysisType")),
            ("Unité", project.get("unit")),
            ("Valeur de référence", project.get("baseline")),
            ("Date d'exécution", run.get("generatedAt")),
            ("Nombre de tirages", run.get("simulations")),
            ("Graine", run.get("seed")),
            ("Corrélations actives", run.get("correlationsEnabled")),
        ],
    )
    row = _section(summary, row, "Indicateurs statistiques", 2)
    summary_values = _mapping(result.get("summary"))
    _key_values(
        summary,
        row,
        [(LABELS.get(str(key), str(key)), value) for key, value in summary_values.items()],
    )
    _finish(summary)

    tables = [
        (
            "Percentiles",
            "Table des percentiles",
            "Niveaux de confiance, valeurs et réserves recommandées",
            "percentiles",
        ),
        (
            "Sensibilité",
            "Analyse de sensibilité",
            "Classement des postes selon leur influence sur le résultat total",
            "sensitivity",
        ),
        (
            "Convergence",
            "Diagnostic de convergence",
            "Stabilité du percentile contrôlé selon le nombre de tirages",
            "convergence",
        ),
    ]
    for sheet_name, title, subtitle, key in tables:
        records = _records(result.get(key))
        sheet = workbook.create_sheet(sheet_name)
        row = _title(sheet, title, subtitle, max(len(records[0]) if records else 2, 2))
        _record_table(sheet, row, records)
        _finish(sheet)

    assumptions = workbook.create_sheet("Hypothèses")
    risk_records = _records(register.get("items"))
    row = _title(
        assumptions,
        "Hypothèses et paramètres",
        "Configuration exacte utilisée pour reproduire la simulation",
        max(len(risk_records[0]) if risk_records else 2, 2),
    )
    row = _section(assumptions, row, "Paramètres de la simulation", 2)
    row = _key_values(
        assumptions,
        row,
        [
            ("Nombre de tirages", config.get("simulations")),
            ("Graine", config.get("seed")),
            (
                "Niveaux de confiance",
                ", ".join(map(str, config.get("levels", [])))
                if isinstance(config.get("levels"), list)
                else config.get("levels"),
            ),
            ("Percentile de décision", config.get("decisionPercentile")),
            ("Seuil de dépassement", config.get("exceedanceThreshold")),
            ("Tolérance de convergence", config.get("convergenceTolerance")),
            ("Nom du scénario", config.get("scenarioName")),
            ("Description du scénario", config.get("scenarioDescription")),
        ],
    )
    row = _section(assumptions, row, "Métadonnées du projet", 2)
    row = _key_values(assumptions, row, [(str(key), value) for key, value in metadata.items()])
    row = _section(
        assumptions, row, "Postes du registre", max(len(risk_records[0]) if risk_records else 2, 2)
    )
    _record_table(assumptions, row, risk_records)
    _finish(assumptions)

    robustness = workbook.create_sheet("Robustesse")
    baseline = _records(result.get("baselineComparison"))
    diagnostics = _records(result.get("correlationDiagnostics"))
    correlation = _mapping(register.get("correlations"))
    matrix_names = _list(correlation.get("names"))
    matrix_values = _list(correlation.get("values"))
    width = max(
        len(baseline[0]) if baseline else 2,
        len(diagnostics[0]) if diagnostics else 2,
        len(matrix_names) + 1,
    )
    row = _title(
        robustness,
        "Robustesse et dépendances",
        "Comparaison à la référence, diagnostics et matrice de corrélation",
        width,
    )
    row = _section(robustness, row, "Comparaison à la référence", width)
    row = _record_table(robustness, row, baseline)
    row = _section(robustness, row, "Diagnostic des corrélations", width)
    row = _record_table(robustness, row, diagnostics)
    row = _section(robustness, row, "Matrice de corrélation", width)
    if matrix_names and len(matrix_values) == len(matrix_names):
        robustness.cell(row=row, column=1, value="Poste").fill = PatternFill(
            "solid", fgColor=ORANGE
        )
        robustness.cell(row=row, column=1).font = Font(bold=True, color=WHITE)
        for column, name in enumerate(matrix_names, start=2):
            cell = robustness.cell(row=row, column=column, value=name)
            cell.fill = PatternFill("solid", fgColor=ORANGE)
            cell.font = Font(bold=True, color=WHITE)
        for row_index, name in enumerate(matrix_names, start=row + 1):
            robustness.cell(row=row_index, column=1, value=name).font = Font(bold=True, color=NAVY)
            for column, value in enumerate(matrix_values[row_index - row - 1], start=2):
                cell = robustness.cell(row=row_index, column=column, value=value)
                cell.number_format = "0.000"
                cell.fill = PatternFill(
                    "solid", fgColor=LIGHT_BLUE if row_index - row == column - 1 else WHITE
                )
    else:
        robustness.cell(
            row=row, column=1, value="Les postes sont considérés comme indépendants."
        ).font = Font(italic=True, color=MUTED)
    _finish(robustness)

    workbook.properties.title = "Résultats de simulation Monte-Carlo"
    workbook.properties.subject = "Export complet et reproductible de la simulation"
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
