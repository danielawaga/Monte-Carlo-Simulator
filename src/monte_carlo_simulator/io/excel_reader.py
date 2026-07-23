"""Read, validate and convert versioned Excel risk registers."""

from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

import numpy as np
import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from monte_carlo_simulator.exceptions import (
    RiskRegisterIssue,
    RiskRegisterValidationError,
    ValidationError,
)
from monte_carlo_simulator.io.schema import (
    CORRELATIONS_SHEET,
    METADATA_SHEET,
    OPTIONAL_RISK_REGISTER_COLUMNS,
    REQUIRED_RISK_REGISTER_COLUMNS,
    RISK_REGISTER_COLUMNS,
    RISK_REGISTER_SHEET,
)
from monte_carlo_simulator.io.validators import (
    MetadataCells,
    dataframe_to_validated_risk_items,
    validate_metadata,
)
from monte_carlo_simulator.models import RiskItem, RiskRegister, RiskRegisterMetadata
from monte_carlo_simulator.models.correlation_matrix import (
    CorrelationMatrix,
    as_correlation_value,
)


def load_workbook(file_path: str | Path) -> Workbook:
    """Open an .xlsx workbook and translate expected file errors for end users."""
    path = Path(file_path)
    if not path.exists():
        raise _file_error(path, "The Excel file does not exist.")
    if not path.is_file():
        raise _file_error(path, "The supplied path is not a file.")
    if path.suffix.casefold() != ".xlsx":
        raise _file_error(path, "Schema v1 requires an .xlsx file.")

    try:
        return openpyxl_load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, ParseError, ValueError) as exc:
        raise _file_error(path, f"The file is not a valid readable .xlsx workbook: {exc}") from exc
    except OSError as exc:
        raise _file_error(path, f"The workbook could not be read: {exc}") from exc


def load_metadata(workbook: Workbook) -> RiskRegisterMetadata:
    """Load and validate the metadata sheet of an open workbook."""
    cells, extraction_issues = _extract_metadata_cells(workbook)
    metadata, validation_issues = validate_metadata(cells)
    issues = [*extraction_issues, *validation_issues]
    if issues:
        raise RiskRegisterValidationError(issues)
    if metadata is None:
        raise RuntimeError("Metadata validation succeeded without producing metadata.")
    return metadata


def load_risk_register_dataframe(workbook: Workbook) -> pd.DataFrame:
    """Load risk_register values and retain their original Excel row numbers."""
    dataframe, issues = _extract_risk_register_dataframe(workbook)
    if issues:
        raise RiskRegisterValidationError(issues)
    return dataframe


def dataframe_to_risk_items(
    dataframe: pd.DataFrame,
    metadata: RiskRegisterMetadata,
) -> list[RiskItem]:
    """Validate active DataFrame rows and convert them to RiskItem objects."""
    items, _source_rows, issues = dataframe_to_validated_risk_items(
        dataframe,
        metadata.default_unit,
    )
    if issues:
        raise RiskRegisterValidationError(issues)
    return items


def load_risk_register(file_path: str | Path) -> RiskRegister:
    """Load the complete workbook and return a validated structured risk register."""
    path = Path(file_path)
    workbook = load_workbook(path)
    try:
        metadata_cells, metadata_extraction_issues = _extract_metadata_cells(workbook)
        metadata, metadata_validation_issues = validate_metadata(metadata_cells)
        dataframe, dataframe_issues = _extract_risk_register_dataframe(workbook)
        provisional_unit = (
            metadata.default_unit if metadata else _provisional_default_unit(metadata_cells)
        )
        items, source_rows, row_issues = dataframe_to_validated_risk_items(
            dataframe,
            provisional_unit,
        )
        correlation_matrix, correlation_issues = _extract_correlation_matrix(workbook, items)
        issues = [
            *metadata_extraction_issues,
            *metadata_validation_issues,
            *dataframe_issues,
            *row_issues,
            *correlation_issues,
        ]
        if issues:
            raise RiskRegisterValidationError(issues)
        if metadata is None:
            raise RuntimeError("Workbook validation succeeded without producing metadata.")
        return RiskRegister(
            metadata=metadata,
            items=items,
            source_path=path.resolve(),
            source_rows=source_rows,
            correlation_matrix=correlation_matrix,
        )
    finally:
        workbook.close()


def load_risk_register_excel(file_path: str | Path) -> pd.DataFrame:
    """Compatibility wrapper returning only the validated risk-register DataFrame."""
    workbook = load_workbook(file_path)
    try:
        return load_risk_register_dataframe(workbook)
    finally:
        workbook.close()


def _extract_metadata_cells(
    workbook: Workbook,
) -> tuple[dict[str, tuple[object, int]], list[RiskRegisterIssue]]:
    issues: list[RiskRegisterIssue] = []
    if METADATA_SHEET not in workbook.sheetnames:
        issues.append(
            RiskRegisterIssue(
                sheet=METADATA_SHEET,
                field="sheet",
                value=METADATA_SHEET,
                message=f"Required worksheet '{METADATA_SHEET}' is missing.",
            )
        )
        return {}, issues

    worksheet = workbook[METADATA_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    header_values = tuple(header or ())
    if len(header_values) < 2 or _normalized_header(header_values[0]) != "key":
        issues.append(
            RiskRegisterIssue(
                sheet=METADATA_SHEET,
                row=1,
                field="key",
                value=header_values[0] if header_values else None,
                message="Cell A1 must contain the header 'key'.",
            )
        )
    if len(header_values) < 2 or _normalized_header(header_values[1]) != "value":
        issues.append(
            RiskRegisterIssue(
                sheet=METADATA_SHEET,
                row=1,
                field="value",
                value=header_values[1] if len(header_values) > 1 else None,
                message="Cell B1 must contain the header 'value'.",
            )
        )

    cells: dict[str, tuple[object, int]] = {}
    for row_number, row in enumerate(rows, start=2):
        values = tuple(row)
        key_value = values[0] if values else None
        cell_value = values[1] if len(values) > 1 else None
        if _is_blank(key_value) and _is_blank(cell_value):
            continue
        if not isinstance(key_value, str) or not key_value.strip():
            issues.append(
                RiskRegisterIssue(
                    sheet=METADATA_SHEET,
                    row=row_number,
                    field="key",
                    value=key_value,
                    message="Metadata keys must be non-empty text.",
                )
            )
            continue
        key = key_value.strip().casefold()
        if key in cells:
            issues.append(
                RiskRegisterIssue(
                    sheet=METADATA_SHEET,
                    row=row_number,
                    field="key",
                    value=key_value,
                    message=f"Duplicate metadata key; first occurrence is on row {cells[key][1]}.",
                )
            )
            continue
        cells[key] = (cell_value, row_number)
    return cells, issues


def _extract_risk_register_dataframe(
    workbook: Workbook,
) -> tuple[pd.DataFrame, list[RiskRegisterIssue]]:
    issues: list[RiskRegisterIssue] = []
    empty = pd.DataFrame(columns=[*RISK_REGISTER_COLUMNS, "_excel_row"])
    if RISK_REGISTER_SHEET not in workbook.sheetnames:
        issues.append(
            RiskRegisterIssue(
                sheet=RISK_REGISTER_SHEET,
                field="sheet",
                value=RISK_REGISTER_SHEET,
                message=f"Required worksheet '{RISK_REGISTER_SHEET}' is missing.",
            )
        )
        return empty, issues

    worksheet = workbook[RISK_REGISTER_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        issues.append(
            RiskRegisterIssue(
                sheet=RISK_REGISTER_SHEET,
                row=1,
                field="columns",
                value=None,
                message="The risk_register worksheet is empty.",
            )
        )
        return empty, issues

    positions: dict[str, int] = {}
    for position, raw_header in enumerate(header):
        if _is_blank(raw_header):
            continue
        if not isinstance(raw_header, str):
            issues.append(
                RiskRegisterIssue(
                    sheet=RISK_REGISTER_SHEET,
                    row=1,
                    field="columns",
                    value=raw_header,
                    message="Column headers must be text.",
                )
            )
            continue
        column = raw_header.strip().casefold()
        if column in positions:
            issues.append(
                RiskRegisterIssue(
                    sheet=RISK_REGISTER_SHEET,
                    row=1,
                    field=column,
                    value=raw_header,
                    message="Duplicate column header.",
                )
            )
            continue
        positions[column] = position

    for column in REQUIRED_RISK_REGISTER_COLUMNS:
        if column not in positions:
            issues.append(
                RiskRegisterIssue(
                    sheet=RISK_REGISTER_SHEET,
                    row=1,
                    field=column,
                    value=None,
                    message=f"Required column '{column}' is missing.",
                )
            )

    records: list[dict[str, object]] = []
    for row_number, row in enumerate(rows, start=2):
        values = tuple(row)
        record = {
            column: values[position] if position < len(values) else None
            for column, position in positions.items()
            if column in RISK_REGISTER_COLUMNS
        }
        if all(_is_blank(record.get(column)) for column in RISK_REGISTER_COLUMNS):
            continue
        for column in RISK_REGISTER_COLUMNS:
            if column not in record:
                if column == "enabled" and column in OPTIONAL_RISK_REGISTER_COLUMNS:
                    record[column] = True
                else:
                    record[column] = None
        record["_excel_row"] = row_number
        records.append(record)

    return pd.DataFrame(records, columns=[*RISK_REGISTER_COLUMNS, "_excel_row"]), issues


def _cell_name(col_index: int, row_number: int) -> str:
    return f"{get_column_letter(col_index + 1)}{row_number}"


def _extract_correlation_matrix(
    workbook: Workbook, active_items: list[RiskItem]
) -> tuple[CorrelationMatrix | None, list[RiskRegisterIssue]]:
    issues: list[RiskRegisterIssue] = []
    if CORRELATIONS_SHEET not in workbook.sheetnames:
        return None, issues
    active_names = tuple(item.name for item in active_items)
    active_lookup = {name.casefold(): name for name in active_names}
    worksheet = workbook[CORRELATIONS_SHEET]
    all_rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    while all_rows and all(_is_blank(cell) for cell in all_rows[-1]):
        all_rows.pop()
    if len(all_rows) < 2:
        issue = RiskRegisterIssue(
            sheet=CORRELATIONS_SHEET,
            row=1,
            field="matrix",
            value=None,
            message="Correlation sheet must contain a header row and one row per active item.",
        )
        return None, [issue]

    last_col = max(len(row) for row in all_rows)
    while last_col > 1:
        column_empty = all(
            col >= len(row) or _is_blank(row[col]) for col in (last_col - 1,) for row in all_rows
        )
        if not column_empty:
            break
        last_col -= 1

    column_names: list[str] = []
    seen_columns: dict[str, int] = {}
    header = all_rows[0]
    for col_index in range(1, last_col):
        value = header[col_index] if col_index < len(header) else None
        cell = _cell_name(col_index, 1)
        if not isinstance(value, str) or not value.strip():
            issues.append(
                RiskRegisterIssue(
                    sheet=CORRELATIONS_SHEET,
                    row=1,
                    field=cell,
                    value=value,
                    message=(
                        "Correlation column names must be non-empty text matching active items."
                    ),
                )
            )
            continue
        key = value.strip().casefold()
        if key in seen_columns:
            issues.append(
                RiskRegisterIssue(
                    sheet=CORRELATIONS_SHEET,
                    row=1,
                    field=cell,
                    value=value,
                    message="Duplicate correlation column name; remove the repeated header.",
                )
            )
            continue
        if key not in active_lookup:
            issues.append(
                RiskRegisterIssue(
                    sheet=CORRELATIONS_SHEET,
                    row=1,
                    field=cell,
                    value=value,
                    message="Unknown correlation column name; use active risk item names.",
                )
            )
            continue
        seen_columns[key] = col_index + 1
        column_names.append(active_lookup[key])

    data_rows = all_rows[1:]
    while data_rows and all(_is_blank(cell) for cell in data_rows[-1]):
        data_rows.pop()
    row_names: list[str] = []
    seen_rows: dict[str, int] = {}
    values: list[list[float]] = []
    for row_number, row in enumerate(data_rows, start=2):
        row_name_value = row[0] if row else None
        has_coefficients = any(
            not _is_blank(row[col] if col < len(row) else None) for col in range(1, last_col)
        )
        if not isinstance(row_name_value, str) or not row_name_value.strip():
            if has_coefficients or not _is_blank(row_name_value):
                issues.append(
                    RiskRegisterIssue(
                        sheet=CORRELATIONS_SHEET,
                        row=row_number,
                        field="row_name",
                        value=row_name_value,
                        message=(
                            "Correlation row names must be non-empty text matching active items."
                        ),
                    )
                )
            continue
        key = row_name_value.strip().casefold()
        if key in seen_rows:
            issues.append(
                RiskRegisterIssue(
                    sheet=CORRELATIONS_SHEET,
                    row=row_number,
                    field="row_name",
                    value=row_name_value,
                    message="Duplicate correlation row name; remove the repeated row.",
                )
            )
            continue
        if key not in active_lookup:
            issues.append(
                RiskRegisterIssue(
                    sheet=CORRELATIONS_SHEET,
                    row=row_number,
                    field="row_name",
                    value=row_name_value,
                    message="Unknown correlation row name; use active risk item names.",
                )
            )
            continue
        seen_rows[key] = row_number
        row_names.append(active_lookup[key])
        coeffs: list[float] = []
        for col_index in range(1, last_col):
            raw = row[col_index] if col_index < len(row) else None
            cell = _cell_name(col_index, row_number)
            if _is_blank(raw):
                issues.append(
                    RiskRegisterIssue(
                        sheet=CORRELATIONS_SHEET,
                        row=row_number,
                        field=cell,
                        value=raw,
                        message="Correlation matrix contains an empty internal coefficient.",
                    )
                )
                coeffs.append(np.nan)
                continue
            try:
                coeffs.append(as_correlation_value(raw))
            except ValidationError:
                issues.append(
                    RiskRegisterIssue(
                        sheet=CORRELATIONS_SHEET,
                        row=row_number,
                        field=cell,
                        value=raw,
                        message="Correlation coefficients must be finite real numbers in [-1, 1].",
                    )
                )
                coeffs.append(np.nan)
        values.append(coeffs)

    if set(row_names) != set(active_names):
        issues.append(
            RiskRegisterIssue(
                sheet=CORRELATIONS_SHEET,
                field="row_names",
                value=row_names,
                message="Correlation row names must match exactly the active risk item names.",
            )
        )
    if set(column_names) != set(active_names):
        issues.append(
            RiskRegisterIssue(
                sheet=CORRELATIONS_SHEET,
                row=1,
                field="column_names",
                value=column_names,
                message="Correlation column names must match exactly the active risk item names.",
            )
        )
    if len(row_names) != len(column_names):
        issues.append(
            RiskRegisterIssue(
                sheet=CORRELATIONS_SHEET,
                field="matrix",
                value=f"{len(row_names)}x{len(column_names)}",
                message="Correlation matrix must be square.",
            )
        )
    if issues:
        return None, issues
    raw_values = np.asarray(values, dtype=float)
    row_order = [row_names.index(name) for name in active_names]
    column_order = [column_names.index(name) for name in active_names]
    aligned = raw_values[np.ix_(row_order, column_order)]
    try:
        return CorrelationMatrix(active_names, aligned), []
    except ValidationError as exc:
        return None, [
            RiskRegisterIssue(
                sheet=CORRELATIONS_SHEET,
                field="matrix",
                value=aligned.tolist(),
                message=f"Invalid correlation matrix: {exc}",
            )
        ]


def _provisional_default_unit(cells: MetadataCells) -> str | None:
    value, _row = cells.get("default_unit", (None, 5))
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def _normalized_header(value: object) -> str | None:
    if isinstance(value, str):
        return value.strip().casefold()
    return None


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return bool(pd.isna(value))


def _file_error(path: Path, message: str) -> RiskRegisterValidationError:
    return RiskRegisterValidationError(
        [
            RiskRegisterIssue(
                sheet="<workbook>",
                field="file_path",
                value=str(path),
                message=message,
            )
        ]
    )
